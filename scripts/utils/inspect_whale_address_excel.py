#!/usr/bin/env python3
"""
whale_address.xlsx 파일 구조 확인 스크립트
"""

import pandas as pd
from openpyxl import load_workbook

excel_file = "whale_address.xlsx"

print("=" * 70)
print("📊 whale_address.xlsx 파일 구조 확인")
print("=" * 70)

try:
    # Excel 파일 로드
    wb = load_workbook(excel_file)
    sheet_names = wb.sheetnames
    
    print(f"\n✅ 파일 로드 성공")
    print(f"📋 총 {len(sheet_names)}개의 탭 발견:")
    for i, sheet_name in enumerate(sheet_names, 1):
        print(f"   {i}. {sheet_name}")
    
    print("\n" + "=" * 70)
    print("각 탭별 데이터 구조 확인")
    print("=" * 70)
    
    for sheet_name in sheet_names:
        print(f"\n📋 탭: {sheet_name}")
        print("-" * 70)
        
        # 탭 읽기
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        
        print(f"   행 수: {len(df)}건")
        print(f"   컬럼 수: {len(df.columns)}개")
        print(f"   컬럼 목록: {list(df.columns)}")
        
        # 샘플 데이터 (상위 3건)
        if len(df) > 0:
            print(f"\n   샘플 데이터 (상위 3건):")
            print(df.head(3).to_string(index=False))
        
        # 데이터 타입 확인
        print(f"\n   데이터 타입:")
        for col in df.columns:
            print(f"      {col}: {df[col].dtype}")
        
        # NULL 값 확인
        print(f"\n   NULL 값 개수:")
        null_counts = df.isnull().sum()
        for col, count in null_counts.items():
            if count > 0:
                print(f"      {col}: {count}개")

except FileNotFoundError:
    print(f"❌ 파일을 찾을 수 없습니다: {excel_file}")
except Exception as e:
    print(f"❌ 오류 발생: {e}")
    import traceback
    traceback.print_exc()



