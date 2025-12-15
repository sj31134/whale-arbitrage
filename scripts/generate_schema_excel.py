#!/usr/bin/env python3
"""
SQL DB와 Supabase 테이블 스키마를 엑셀 파일로 생성
각 테이블마다 하나의 탭으로 구성하고 스키마 구조를 설명
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import sys

# 프로젝트 루트 경로
PROJECT_ROOT = Path(__file__).parent.parent

# 테이블 스키마 정의
TABLE_SCHEMAS = {
    # Supabase 테이블들 (create_project_tables.sql 기반)
    "binance_futures_metrics": {
        "description": "바이낸스 선물 지표 데이터",
        "database": "Supabase",
        "columns": [
            {"name": "id", "type": "SERIAL", "nullable": False, "description": "기본 키"},
            {"name": "date", "type": "DATE", "nullable": False, "description": "날짜"},
            {"name": "symbol", "type": "VARCHAR(20)", "nullable": False, "description": "심볼 (예: BTCUSDT)"},
            {"name": "avg_funding_rate", "type": "DECIMAL(20,10)", "nullable": True, "description": "평균 펀딩 레이트"},
            {"name": "sum_open_interest", "type": "DECIMAL(30,10)", "nullable": True, "description": "총 미체결약정"},
            {"name": "long_short_ratio", "type": "DECIMAL(10,6)", "nullable": True, "description": "롱/숏 비율"},
            {"name": "volatility_24h", "type": "DECIMAL(10,6)", "nullable": True, "description": "24시간 변동성"},
            {"name": "target_volatility_24h", "type": "DECIMAL(10,6)", "nullable": True, "description": "목표 변동성 24시간"},
            {"name": "created_at", "type": "TIMESTAMP", "nullable": True, "description": "생성 시간"},
        ],
        "constraints": ["PRIMARY KEY (id)", "UNIQUE(date, symbol)"],
        "indexes": ["idx_futures_date", "idx_futures_symbol"]
    },
    
    "binance_spot_daily": {
        "description": "바이낸스 현물 일일 가격 데이터",
        "database": "Supabase",
        "columns": [
            {"name": "id", "type": "SERIAL", "nullable": False, "description": "기본 키"},
            {"name": "symbol", "type": "VARCHAR(20)", "nullable": False, "description": "심볼"},
            {"name": "date", "type": "DATE", "nullable": False, "description": "날짜"},
            {"name": "open", "type": "DECIMAL(20,8)", "nullable": True, "description": "시가"},
            {"name": "high", "type": "DECIMAL(20,8)", "nullable": True, "description": "고가"},
            {"name": "low", "type": "DECIMAL(20,8)", "nullable": True, "description": "저가"},
            {"name": "close", "type": "DECIMAL(20,8)", "nullable": True, "description": "종가"},
            {"name": "volume", "type": "DECIMAL(30,8)", "nullable": True, "description": "거래량"},
            {"name": "quote_volume", "type": "DECIMAL(30,8)", "nullable": True, "description": "기준 통화 거래량"},
            {"name": "created_at", "type": "TIMESTAMP", "nullable": True, "description": "생성 시간"},
        ],
        "constraints": ["PRIMARY KEY (id)", "UNIQUE(date, symbol)"],
        "indexes": ["idx_spot_daily_date", "idx_spot_daily_symbol"]
    },
    
    "binance_spot_weekly": {
        "description": "바이낸스 현물 주간 가격 및 기술적 지표 데이터",
        "database": "Supabase",
        "columns": [
            {"name": "id", "type": "SERIAL", "nullable": False, "description": "기본 키"},
            {"name": "symbol", "type": "VARCHAR(20)", "nullable": False, "description": "심볼"},
            {"name": "date", "type": "DATE", "nullable": False, "description": "날짜"},
            {"name": "open", "type": "DECIMAL(20,8)", "nullable": True, "description": "시가"},
            {"name": "high", "type": "DECIMAL(20,8)", "nullable": True, "description": "고가"},
            {"name": "low", "type": "DECIMAL(20,8)", "nullable": True, "description": "저가"},
            {"name": "close", "type": "DECIMAL(20,8)", "nullable": True, "description": "종가"},
            {"name": "volume", "type": "DECIMAL(30,8)", "nullable": True, "description": "거래량"},
            {"name": "quote_volume", "type": "DECIMAL(30,8)", "nullable": True, "description": "기준 통화 거래량"},
            {"name": "atr", "type": "DECIMAL(20,8)", "nullable": True, "description": "ATR (Average True Range)"},
            {"name": "rsi", "type": "DECIMAL(10,4)", "nullable": True, "description": "RSI (Relative Strength Index)"},
            {"name": "upper_shadow", "type": "DECIMAL(20,8)", "nullable": True, "description": "위 그림자"},
            {"name": "lower_shadow", "type": "DECIMAL(20,8)", "nullable": True, "description": "아래 그림자"},
            {"name": "upper_shadow_ratio", "type": "DECIMAL(10,6)", "nullable": True, "description": "위 그림자 비율"},
            {"name": "lower_shadow_ratio", "type": "DECIMAL(10,6)", "nullable": True, "description": "아래 그림자 비율"},
            {"name": "weekly_range", "type": "DECIMAL(20,8)", "nullable": True, "description": "주간 범위"},
            {"name": "weekly_range_pct", "type": "DECIMAL(10,6)", "nullable": True, "description": "주간 범위 비율(%)"},
            {"name": "body_size", "type": "DECIMAL(20,8)", "nullable": True, "description": "몸통 크기"},
            {"name": "body_size_pct", "type": "DECIMAL(10,6)", "nullable": True, "description": "몸통 크기 비율(%)"},
            {"name": "volatility_ratio", "type": "DECIMAL(10,6)", "nullable": True, "description": "변동성 비율"},
            {"name": "created_at", "type": "TIMESTAMP", "nullable": True, "description": "생성 시간"},
        ],
        "constraints": ["PRIMARY KEY (id)", "UNIQUE(date, symbol)"],
        "indexes": ["idx_spot_weekly_date", "idx_spot_weekly_symbol"]
    },
    
    "bitget_spot_daily": {
        "description": "비트겟 현물 일일 가격 데이터",
        "database": "Supabase",
        "columns": [
            {"name": "id", "type": "SERIAL", "nullable": False, "description": "기본 키"},
            {"name": "symbol", "type": "VARCHAR(20)", "nullable": False, "description": "심볼"},
            {"name": "date", "type": "DATE", "nullable": False, "description": "날짜"},
            {"name": "open", "type": "DECIMAL(20,8)", "nullable": True, "description": "시가"},
            {"name": "high", "type": "DECIMAL(20,8)", "nullable": True, "description": "고가"},
            {"name": "low", "type": "DECIMAL(20,8)", "nullable": True, "description": "저가"},
            {"name": "close", "type": "DECIMAL(20,8)", "nullable": True, "description": "종가"},
            {"name": "volume", "type": "DECIMAL(30,8)", "nullable": True, "description": "거래량"},
            {"name": "quote_volume", "type": "DECIMAL(30,8)", "nullable": True, "description": "기준 통화 거래량"},
            {"name": "created_at", "type": "TIMESTAMP", "nullable": True, "description": "생성 시간"},
        ],
        "constraints": ["PRIMARY KEY (id)", "UNIQUE(date, symbol)"],
        "indexes": ["idx_bitget_date", "idx_bitget_symbol"]
    },
    
    "bybit_spot_daily": {
        "description": "바이빗 현물 일일 가격 데이터",
        "database": "Supabase",
        "columns": [
            {"name": "id", "type": "SERIAL", "nullable": False, "description": "기본 키"},
            {"name": "symbol", "type": "VARCHAR(20)", "nullable": False, "description": "심볼"},
            {"name": "date", "type": "DATE", "nullable": False, "description": "날짜"},
            {"name": "open", "type": "DECIMAL(20,8)", "nullable": True, "description": "시가"},
            {"name": "high", "type": "DECIMAL(20,8)", "nullable": True, "description": "고가"},
            {"name": "low", "type": "DECIMAL(20,8)", "nullable": True, "description": "저가"},
            {"name": "close", "type": "DECIMAL(20,8)", "nullable": True, "description": "종가"},
            {"name": "volume", "type": "DECIMAL(30,8)", "nullable": True, "description": "거래량"},
            {"name": "quote_volume", "type": "DECIMAL(30,8)", "nullable": True, "description": "기준 통화 거래량"},
            {"name": "created_at", "type": "TIMESTAMP", "nullable": True, "description": "생성 시간"},
        ],
        "constraints": ["PRIMARY KEY (id)", "UNIQUE(date, symbol)"],
        "indexes": ["idx_bybit_spot_date", "idx_bybit_spot_symbol"]
    },
    
    "upbit_daily": {
        "description": "업비트 일일 가격 데이터",
        "database": "Supabase",
        "columns": [
            {"name": "id", "type": "SERIAL", "nullable": False, "description": "기본 키"},
            {"name": "market", "type": "VARCHAR(20)", "nullable": False, "description": "마켓 (예: KRW-BTC)"},
            {"name": "date", "type": "DATE", "nullable": False, "description": "날짜"},
            {"name": "opening_price", "type": "DECIMAL(20,8)", "nullable": True, "description": "시가"},
            {"name": "high_price", "type": "DECIMAL(20,8)", "nullable": True, "description": "고가"},
            {"name": "low_price", "type": "DECIMAL(20,8)", "nullable": True, "description": "저가"},
            {"name": "trade_price", "type": "DECIMAL(20,8)", "nullable": True, "description": "종가"},
            {"name": "acc_trade_volume_24h", "type": "DECIMAL(30,8)", "nullable": True, "description": "24시간 누적 거래량"},
            {"name": "acc_trade_price_24h", "type": "DECIMAL(30,8)", "nullable": True, "description": "24시간 누적 거래금액"},
            {"name": "created_at", "type": "TIMESTAMP", "nullable": True, "description": "생성 시간"},
        ],
        "constraints": ["PRIMARY KEY (id)", "UNIQUE(date, market)"],
        "indexes": ["idx_upbit_date", "idx_upbit_market"]
    },
    
    "bitinfocharts_whale": {
        "description": "비트인포차트 고래 데이터 (일별)",
        "database": "Supabase",
        "columns": [
            {"name": "id", "type": "SERIAL", "nullable": False, "description": "기본 키"},
            {"name": "date", "type": "DATE", "nullable": False, "description": "날짜"},
            {"name": "coin", "type": "VARCHAR(10)", "nullable": False, "description": "코인 심볼"},
            {"name": "top100_richest_pct", "type": "DECIMAL(10,4)", "nullable": True, "description": "상위 100명 보유 비율(%)"},
            {"name": "avg_transaction_value_btc", "type": "DECIMAL(20,8)", "nullable": True, "description": "평균 거래 가치 (BTC)"},
            {"name": "top10_pct", "type": "DECIMAL(10,4)", "nullable": True, "description": "상위 10명 보유 비율(%)"},
            {"name": "created_at", "type": "TIMESTAMP", "nullable": True, "description": "생성 시간"},
        ],
        "constraints": ["PRIMARY KEY (id)", "UNIQUE(date, coin)"],
        "indexes": ["idx_whale_date", "idx_whale_coin"]
    },
    
    "bitinfocharts_whale_weekly": {
        "description": "비트인포차트 고래 데이터 (주별)",
        "database": "Supabase",
        "columns": [
            {"name": "coin", "type": "VARCHAR(10)", "nullable": False, "description": "코인 심볼"},
            {"name": "week_end_date", "type": "DATE", "nullable": False, "description": "주말 날짜"},
            {"name": "avg_top100_richest_pct", "type": "DECIMAL(10,4)", "nullable": True, "description": "평균 상위 100명 보유 비율(%)"},
            {"name": "avg_transaction_value_btc", "type": "DECIMAL(20,8)", "nullable": True, "description": "평균 거래 가치 (BTC)"},
            {"name": "whale_conc_change_7d", "type": "DECIMAL(10,6)", "nullable": True, "description": "7일 고래 집중도 변화"},
            {"name": "created_at", "type": "TIMESTAMP", "nullable": True, "description": "생성 시간"},
        ],
        "constraints": ["PRIMARY KEY (coin, week_end_date)"],
        "indexes": []
    },
    
    "whale_weekly_stats": {
        "description": "고래 주간 통계 데이터",
        "database": "Supabase",
        "columns": [
            {"name": "id", "type": "SERIAL", "nullable": False, "description": "기본 키"},
            {"name": "date", "type": "DATE", "nullable": False, "description": "날짜"},
            {"name": "coin_symbol", "type": "VARCHAR(20)", "nullable": False, "description": "코인 심볼"},
            {"name": "net_inflow_usd", "type": "DECIMAL(30,8)", "nullable": True, "description": "순 유입 (USD)"},
            {"name": "exchange_inflow_usd", "type": "DECIMAL(30,8)", "nullable": True, "description": "거래소 유입 (USD)"},
            {"name": "active_addresses", "type": "INTEGER", "nullable": True, "description": "활성 주소 수"},
            {"name": "transaction_count", "type": "INTEGER", "nullable": True, "description": "거래 건수"},
            {"name": "avg_buy_price", "type": "DECIMAL(20,8)", "nullable": True, "description": "평균 매수 가격"},
            {"name": "created_at", "type": "TIMESTAMP", "nullable": True, "description": "생성 시간"},
        ],
        "constraints": ["PRIMARY KEY (id)", "UNIQUE(date, coin_symbol)"],
        "indexes": ["idx_whale_weekly_date", "idx_whale_weekly_coin"]
    },
    
    "whale_daily_stats": {
        "description": "고래 일일 통계 데이터",
        "database": "Supabase",
        "columns": [
            {"name": "id", "type": "SERIAL", "nullable": False, "description": "기본 키"},
            {"name": "date", "type": "DATE", "nullable": False, "description": "날짜"},
            {"name": "coin_symbol", "type": "VARCHAR(20)", "nullable": False, "description": "코인 심볼"},
            {"name": "exchange_inflow_usd", "type": "DECIMAL(30,8)", "nullable": True, "description": "거래소 유입 (USD)"},
            {"name": "exchange_outflow_usd", "type": "DECIMAL(30,8)", "nullable": True, "description": "거래소 유출 (USD)"},
            {"name": "net_flow_usd", "type": "DECIMAL(30,8)", "nullable": True, "description": "순 유동 (USD)"},
            {"name": "whale_to_whale_usd", "type": "DECIMAL(30,8)", "nullable": True, "description": "고래간 거래 (USD)"},
            {"name": "active_addresses", "type": "INTEGER", "nullable": True, "description": "활성 주소 수"},
            {"name": "large_tx_count", "type": "INTEGER", "nullable": True, "description": "대형 거래 건수"},
            {"name": "avg_tx_size_usd", "type": "DECIMAL(20,8)", "nullable": True, "description": "평균 거래 크기 (USD)"},
            {"name": "created_at", "type": "TIMESTAMP", "nullable": True, "description": "생성 시간"},
        ],
        "constraints": ["PRIMARY KEY (id)", "UNIQUE(date, coin_symbol)"],
        "indexes": ["idx_whale_daily_date", "idx_whale_daily_coin"]
    },
    
    "futures_extended_metrics": {
        "description": "선물 확장 지표 데이터 (롱숏비율, Taker비율, Bybit 데이터)",
        "database": "Supabase",
        "columns": [
            {"name": "id", "type": "SERIAL", "nullable": False, "description": "기본 키"},
            {"name": "date", "type": "DATE", "nullable": False, "description": "날짜"},
            {"name": "symbol", "type": "VARCHAR(20)", "nullable": False, "description": "심볼"},
            {"name": "long_short_ratio", "type": "DECIMAL(10,6)", "nullable": True, "description": "롱/숏 비율"},
            {"name": "long_account_pct", "type": "DECIMAL(10,6)", "nullable": True, "description": "롱 계정 비율(%)"},
            {"name": "short_account_pct", "type": "DECIMAL(10,6)", "nullable": True, "description": "숏 계정 비율(%)"},
            {"name": "taker_buy_sell_ratio", "type": "DECIMAL(10,6)", "nullable": True, "description": "Taker 매수/매도 비율"},
            {"name": "taker_buy_vol", "type": "DECIMAL(30,8)", "nullable": True, "description": "Taker 매수 거래량"},
            {"name": "taker_sell_vol", "type": "DECIMAL(30,8)", "nullable": True, "description": "Taker 매도 거래량"},
            {"name": "top_trader_long_short_ratio", "type": "DECIMAL(10,6)", "nullable": True, "description": "상위 트레이더 롱/숏 비율"},
            {"name": "bybit_funding_rate", "type": "DECIMAL(20,10)", "nullable": True, "description": "Bybit 펀딩 레이트"},
            {"name": "bybit_oi", "type": "DECIMAL(30,10)", "nullable": True, "description": "Bybit 미체결약정"},
            {"name": "created_at", "type": "TIMESTAMP", "nullable": True, "description": "생성 시간"},
        ],
        "constraints": ["PRIMARY KEY (id)", "UNIQUE(date, symbol)"],
        "indexes": ["idx_ext_metrics_date", "idx_ext_metrics_symbol"]
    },
    
    "binance_futures_weekly": {
        "description": "바이낸스 선물 주간 지표 데이터",
        "database": "Supabase",
        "columns": [
            {"name": "symbol", "type": "VARCHAR(20)", "nullable": False, "description": "심볼"},
            {"name": "week_end_date", "type": "DATE", "nullable": False, "description": "주말 날짜"},
            {"name": "avg_funding_rate", "type": "DECIMAL(20,10)", "nullable": True, "description": "평균 펀딩 레이트"},
            {"name": "sum_open_interest", "type": "DECIMAL(30,10)", "nullable": True, "description": "총 미체결약정"},
            {"name": "oi_growth_7d", "type": "DECIMAL(10,6)", "nullable": True, "description": "7일 미체결약정 성장률"},
            {"name": "funding_rate_zscore", "type": "DECIMAL(10,6)", "nullable": True, "description": "펀딩 레이트 Z-score"},
            {"name": "created_at", "type": "TIMESTAMP", "nullable": True, "description": "생성 시간"},
        ],
        "constraints": ["PRIMARY KEY (symbol, week_end_date)"],
        "indexes": ["idx_futures_weekly_symbol", "idx_futures_weekly_date"]
    },
    
    "exchange_rate": {
        "description": "환율 데이터 (KRW/USD)",
        "database": "Supabase",
        "columns": [
            {"name": "id", "type": "SERIAL", "nullable": False, "description": "기본 키"},
            {"name": "date", "type": "DATE", "nullable": False, "description": "날짜"},
            {"name": "krw_usd", "type": "DECIMAL(10,4)", "nullable": True, "description": "KRW/USD 환율"},
            {"name": "created_at", "type": "TIMESTAMP", "nullable": True, "description": "생성 시간"},
        ],
        "constraints": ["PRIMARY KEY (id)", "UNIQUE(date)"],
        "indexes": ["idx_exchange_date"]
    },
    
    "internal_transactions": {
        "description": "내부 거래 데이터 (스마트 컨트랙트 내부 거래)",
        "database": "Supabase",
        "columns": [
            {"name": "id", "type": "BIGSERIAL", "nullable": False, "description": "기본 키"},
            {"name": "tx_hash", "type": "TEXT", "nullable": False, "description": "트랜잭션 해시"},
            {"name": "trace_id", "type": "TEXT", "nullable": False, "description": "트레이스 ID"},
            {"name": "block_number", "type": "BIGINT", "nullable": False, "description": "블록 번호"},
            {"name": "block_timestamp", "type": "TIMESTAMPTZ", "nullable": False, "description": "블록 타임스탬프"},
            {"name": "from_address", "type": "TEXT", "nullable": False, "description": "발신 주소"},
            {"name": "to_address", "type": "TEXT", "nullable": True, "description": "수신 주소"},
            {"name": "contract_address", "type": "TEXT", "nullable": True, "description": "컨트랙트 주소"},
            {"name": "value_eth", "type": "NUMERIC(78,18)", "nullable": False, "description": "ETH 가치 (Wei 단위)"},
            {"name": "value_usd", "type": "NUMERIC(20,2)", "nullable": True, "description": "USD 가치"},
            {"name": "transaction_type", "type": "TEXT", "nullable": False, "description": "거래 유형 (CALL, CREATE, SUICIDE 등)"},
            {"name": "is_error", "type": "BOOLEAN", "nullable": False, "description": "에러 여부"},
            {"name": "input_data", "type": "TEXT", "nullable": True, "description": "입력 데이터"},
            {"name": "gas", "type": "BIGINT", "nullable": True, "description": "가스"},
            {"name": "gas_used", "type": "BIGINT", "nullable": True, "description": "사용된 가스"},
            {"name": "created_at", "type": "TIMESTAMPTZ", "nullable": True, "description": "생성 시간"},
            {"name": "updated_at", "type": "TIMESTAMPTZ", "nullable": True, "description": "수정 시간"},
        ],
        "constraints": ["PRIMARY KEY (id)", "UNIQUE(tx_hash, trace_id)"],
        "indexes": ["idx_internal_tx_hash", "idx_internal_tx_from", "idx_internal_tx_to", "idx_internal_tx_contract", "idx_internal_tx_block_number", "idx_internal_tx_timestamp"]
    },
    
    # 기존 문서에서 언급된 테이블들 (일반적인 스키마)
    "cryptocurrencies": {
        "description": "암호화폐 기본 정보",
        "database": "Supabase",
        "columns": [
            {"name": "id", "type": "UUID", "nullable": False, "description": "기본 키"},
            {"name": "symbol", "type": "VARCHAR", "nullable": False, "description": "심볼 (예: BTC, ETH)"},
            {"name": "name", "type": "VARCHAR", "nullable": True, "description": "코인 이름"},
            {"name": "binance_symbol", "type": "VARCHAR", "nullable": True, "description": "바이낸스 심볼"},
            {"name": "market_cap_rank", "type": "INTEGER", "nullable": True, "description": "시가총액 순위"},
            {"name": "is_active", "type": "BOOLEAN", "nullable": True, "description": "활성 여부"},
            {"name": "created_at", "type": "TIMESTAMPTZ", "nullable": True, "description": "생성 시간"},
            {"name": "updated_at", "type": "TIMESTAMPTZ", "nullable": True, "description": "수정 시간"},
        ],
        "constraints": ["PRIMARY KEY (id)", "UNIQUE(symbol)"],
        "indexes": []
    },
    
    "whale_address": {
        "description": "고래 지갑 주소 정보",
        "database": "Supabase",
        "columns": [
            {"name": "id", "type": "TEXT", "nullable": False, "description": "기본 키"},
            {"name": "chain_type", "type": "TEXT", "nullable": False, "description": "체인 타입 (ETH, BSC 등)"},
            {"name": "address", "type": "TEXT", "nullable": False, "description": "지갑 주소"},
            {"name": "name_tag", "type": "TEXT", "nullable": True, "description": "이름 태그"},
            {"name": "balance", "type": "TEXT", "nullable": True, "description": "잔액"},
            {"name": "percentage", "type": "TEXT", "nullable": True, "description": "비율"},
            {"name": "txn_count", "type": "TEXT", "nullable": True, "description": "거래 건수"},
        ],
        "constraints": ["PRIMARY KEY (id, chain_type)"],
        "indexes": []
    },
    
    "whale_transactions": {
        "description": "고래 거래 기록",
        "database": "Supabase",
        "columns": [
            {"name": "tx_hash", "type": "TEXT", "nullable": False, "description": "트랜잭션 해시 (기본 키)"},
            {"name": "block_number", "type": "BIGINT", "nullable": True, "description": "블록 번호"},
            {"name": "block_timestamp", "type": "TIMESTAMPTZ", "nullable": True, "description": "블록 타임스탬프"},
            {"name": "from_address", "type": "TEXT", "nullable": True, "description": "발신 주소"},
            {"name": "to_address", "type": "TEXT", "nullable": True, "description": "수신 주소"},
            {"name": "coin_symbol", "type": "TEXT", "nullable": True, "description": "코인 심볼"},
            {"name": "amount", "type": "NUMERIC", "nullable": True, "description": "거래 금액"},
            {"name": "amount_usd", "type": "NUMERIC", "nullable": True, "description": "거래 금액 (USD)"},
            {"name": "transaction_status", "type": "TEXT", "nullable": True, "description": "거래 상태"},
            {"name": "is_whale", "type": "BOOLEAN", "nullable": True, "description": "고래 여부"},
            {"name": "whale_category", "type": "TEXT", "nullable": True, "description": "고래 카테고리"},
            {"name": "chain", "type": "VARCHAR", "nullable": True, "description": "체인"},
            {"name": "transaction_direction", "type": "TEXT", "nullable": True, "description": "거래 방향 (매수/매도)"},
        ],
        "constraints": ["PRIMARY KEY (tx_hash)"],
        "indexes": []
    },
    
    "influencer": {
        "description": "인플루언서 포스트 데이터",
        "database": "Supabase",
        "columns": [
            {"name": "id", "type": "INTEGER", "nullable": False, "description": "기본 키"},
            {"name": "influencer_id", "type": "VARCHAR", "nullable": True, "description": "인플루언서 ID"},
            {"name": "influencer_name", "type": "VARCHAR", "nullable": True, "description": "인플루언서 이름"},
            {"name": "platform", "type": "VARCHAR", "nullable": True, "description": "플랫폼 (Twitter, Reddit 등)"},
            {"name": "content", "type": "TEXT", "nullable": True, "description": "포스트 내용"},
            {"name": "p_coin_name", "type": "VARCHAR", "nullable": True, "description": "언급된 코인 이름"},
            {"name": "p_sentiment_score", "type": "DOUBLE PRECISION", "nullable": True, "description": "감정 점수"},
            {"name": "retweet_count", "type": "INTEGER", "nullable": True, "description": "리트윗 수"},
            {"name": "engagement", "type": "INTEGER", "nullable": True, "description": "참여도"},
            {"name": "post_date", "type": "TIMESTAMP", "nullable": True, "description": "포스트 날짜"},
        ],
        "constraints": ["PRIMARY KEY (id)"],
        "indexes": []
    },
    
    "price_history": {
        "description": "가격 이력 데이터",
        "database": "Supabase",
        "columns": [
            {"name": "id", "type": "UUID", "nullable": False, "description": "기본 키"},
            {"name": "crypto_id", "type": "UUID", "nullable": True, "description": "암호화폐 ID (FK)"},
            {"name": "timestamp", "type": "TIMESTAMPTZ", "nullable": True, "description": "타임스탬프"},
            {"name": "open_price", "type": "NUMERIC", "nullable": True, "description": "시가"},
            {"name": "high_price", "type": "NUMERIC", "nullable": True, "description": "고가"},
            {"name": "low_price", "type": "NUMERIC", "nullable": True, "description": "저가"},
            {"name": "close_price", "type": "NUMERIC", "nullable": True, "description": "종가"},
            {"name": "volume", "type": "NUMERIC", "nullable": True, "description": "거래량"},
        ],
        "constraints": ["PRIMARY KEY (id)"],
        "indexes": []
    },
    
    "market_cap_data": {
        "description": "시가총액 데이터",
        "database": "Supabase",
        "columns": [
            {"name": "crypto_id", "type": "UUID", "nullable": False, "description": "암호화폐 ID (FK)"},
            {"name": "timestamp", "type": "TIMESTAMPTZ", "nullable": False, "description": "타임스탬프"},
            {"name": "market_cap", "type": "NUMERIC", "nullable": True, "description": "시가총액"},
        ],
        "constraints": ["PRIMARY KEY (crypto_id, timestamp)"],
        "indexes": []
    },
    
    "market_data_daily": {
        "description": "일일 시장 데이터",
        "database": "Supabase",
        "columns": [
            {"name": "crypto_id", "type": "UUID", "nullable": False, "description": "암호화폐 ID (FK)"},
            {"name": "date", "type": "DATE", "nullable": False, "description": "날짜"},
            {"name": "open_price", "type": "NUMERIC", "nullable": True, "description": "시가"},
            {"name": "close_price", "type": "NUMERIC", "nullable": True, "description": "종가"},
        ],
        "constraints": ["PRIMARY KEY (crypto_id, date)"],
        "indexes": []
    },
    
    "reddit_sentiment": {
        "description": "레딧 감정 분석 데이터",
        "database": "Supabase",
        "columns": [
            {"name": "id", "type": "UUID", "nullable": False, "description": "기본 키"},
            {"name": "crypto_id", "type": "UUID", "nullable": True, "description": "암호화폐 ID (FK)"},
            {"name": "timestamp", "type": "TIMESTAMPTZ", "nullable": True, "description": "타임스탬프"},
            {"name": "total_mentions", "type": "INTEGER", "nullable": True, "description": "총 언급 수"},
            {"name": "positive_mentions", "type": "INTEGER", "nullable": True, "description": "긍정 언급 수"},
            {"name": "negative_mentions", "type": "INTEGER", "nullable": True, "description": "부정 언급 수"},
            {"name": "neutral_mentions", "type": "INTEGER", "nullable": True, "description": "중립 언급 수"},
            {"name": "sentiment_score", "type": "NUMERIC", "nullable": True, "description": "감정 점수"},
        ],
        "constraints": ["PRIMARY KEY (id)"],
        "indexes": []
    },
    
    "news_sentiment": {
        "description": "뉴스 감정 분석 데이터",
        "database": "Supabase",
        "columns": [
            {"name": "id", "type": "UUID", "nullable": False, "description": "기본 키"},
            {"name": "crypto_id", "type": "UUID", "nullable": True, "description": "암호화폐 ID (FK)"},
            {"name": "timestamp", "type": "TIMESTAMPTZ", "nullable": True, "description": "타임스탬프"},
            {"name": "sentiment_score", "type": "NUMERIC", "nullable": True, "description": "감정 점수"},
        ],
        "constraints": ["PRIMARY KEY (id)"],
        "indexes": []
    },
    
    "social_data": {
        "description": "소셜 미디어 데이터",
        "database": "Supabase",
        "columns": [
            {"name": "id", "type": "UUID", "nullable": False, "description": "기본 키"},
            {"name": "crypto_id", "type": "UUID", "nullable": True, "description": "암호화폐 ID (FK)"},
            {"name": "timestamp", "type": "TIMESTAMPTZ", "nullable": True, "description": "타임스탬프"},
            {"name": "platform", "type": "VARCHAR", "nullable": True, "description": "플랫폼"},
            {"name": "mention_count", "type": "INTEGER", "nullable": True, "description": "언급 수"},
        ],
        "constraints": ["PRIMARY KEY (id)"],
        "indexes": []
    },
    
    "prediction_accuracy": {
        "description": "예측 정확도 데이터",
        "database": "Supabase",
        "columns": [
            {"name": "id", "type": "UUID", "nullable": False, "description": "기본 키"},
            {"name": "crypto_id", "type": "UUID", "nullable": True, "description": "암호화폐 ID (FK)"},
            {"name": "prediction_date", "type": "DATE", "nullable": True, "description": "예측 날짜"},
            {"name": "accuracy_score", "type": "NUMERIC", "nullable": True, "description": "정확도 점수"},
        ],
        "constraints": ["PRIMARY KEY (id)"],
        "indexes": []
    },
}


def create_excel_file(output_path: Path):
    """엑셀 파일 생성"""
    wb = Workbook()
    
    # 기본 시트 제거
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 스타일 정의
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 각 테이블마다 시트 생성
    for table_name, schema_info in sorted(TABLE_SCHEMAS.items()):
        ws = wb.create_sheet(title=table_name[:31])  # 엑셀 시트 이름은 31자 제한
        
        # 제목 행
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = f"테이블: {table_name}"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 테이블 정보
        row = 3
        ws[f'A{row}'] = "데이터베이스"
        ws[f'B{row}'] = schema_info['database']
        ws[f'A{row+1}'] = "설명"
        ws[f'B{row+1}'] = schema_info['description']
        
        # 컬럼 정보 헤더
        row = 6
        headers = ["컬럼명", "데이터 타입", "NULL 허용", "설명"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # 컬럼 정보 데이터
        for col_info in schema_info['columns']:
            row += 1
            ws.cell(row=row, column=1, value=col_info['name']).border = border
            ws.cell(row=row, column=2, value=col_info['type']).border = border
            ws.cell(row=row, column=3, value="NOT NULL" if not col_info['nullable'] else "NULL").border = border
            ws.cell(row=row, column=4, value=col_info['description']).border = border
        
        # 제약 조건
        row += 2
        ws.cell(row=row, column=1, value="제약 조건").font = Font(bold=True, size=12)
        row += 1
        for constraint in schema_info['constraints']:
            ws.cell(row=row, column=1, value=constraint)
            row += 1
        
        # 인덱스
        if schema_info['indexes']:
            row += 1
            ws.cell(row=row, column=1, value="인덱스").font = Font(bold=True, size=12)
            row += 1
            for index in schema_info['indexes']:
                ws.cell(row=row, column=1, value=index)
                row += 1
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 50
    
    # 목차 시트 생성
    toc_ws = wb.create_sheet(title="목차", index=0)
    toc_ws['A1'] = "코인 데이터베이스 스키마 구조"
    toc_ws['A1'].font = Font(bold=True, size=16)
    
    toc_ws['A3'] = "테이블 목록"
    toc_ws['A3'].font = Font(bold=True, size=12)
    
    headers = ["번호", "테이블명", "데이터베이스", "설명"]
    for col_idx, header in enumerate(headers, 1):
        cell = toc_ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row = 5
    for idx, (table_name, schema_info) in enumerate(sorted(TABLE_SCHEMAS.items()), 1):
        toc_ws.cell(row=row, column=1, value=idx).border = border
        toc_ws.cell(row=row, column=2, value=table_name).border = border
        toc_ws.cell(row=row, column=3, value=schema_info['database']).border = border
        toc_ws.cell(row=row, column=4, value=schema_info['description']).border = border
        row += 1
    
    # 목차 열 너비 조정
    toc_ws.column_dimensions['A'].width = 8
    toc_ws.column_dimensions['B'].width = 30
    toc_ws.column_dimensions['C'].width = 15
    toc_ws.column_dimensions['D'].width = 50
    
    # 파일 저장
    wb.save(output_path)
    print(f"✅ 엑셀 파일 생성 완료: {output_path}")
    print(f"   총 {len(TABLE_SCHEMAS)}개의 테이블 스키마가 포함되었습니다.")


def main():
    """메인 함수"""
    output_path = PROJECT_ROOT / "코인_데이터스키마_구조.xlsx"
    
    print("=" * 60)
    print("📊 코인 데이터베이스 스키마 엑셀 파일 생성")
    print("=" * 60)
    print(f"\n출력 파일: {output_path}")
    print(f"테이블 수: {len(TABLE_SCHEMAS)}개\n")
    
    create_excel_file(output_path)
    
    print("\n✅ 작업 완료!")


if __name__ == "__main__":
    main()


