#!/usr/bin/env python3
"""
Supabase 테이블 스키마를 Excel 파일로 생성하는 스크립트
이미지와 동일한 형식으로 각 테이블을 별도 탭으로 구성
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 테이블 스키마 정의 (CREATE TABLE 문 기반)
table_schemas = {
    "cryptocurrencies": {
        "테이블명": "암호화폐 정보",
        "columns": [
            {"컬럼": "id", "컬럼명": "고유 ID", "데이타타입": "UUID", "PK여부": "Y"},
            {"컬럼": "symbol", "컬럼명": "심볼", "데이타타입": "VARCHAR", "PK여부": ""},
            {"컬럼": "name", "컬럼명": "이름", "데이타타입": "VARCHAR", "PK여부": ""},
            {"컬럼": "binance_symbol", "컬럼명": "바이낸스 심볼", "데이타타입": "VARCHAR", "PK여부": ""},
            {"컬럼": "market_cap_rank", "컬럼명": "시가총액 순위", "데이타타입": "INTEGER", "PK여부": ""},
            {"컬럼": "is_active", "컬럼명": "활성 여부", "데이타타입": "BOOLEAN", "PK여부": ""},
            {"컬럼": "created_at", "컬럼명": "생성일시", "데이타타입": "TIMESTAMPTZ", "PK여부": ""},
            {"컬럼": "updated_at", "컬럼명": "수정일시", "데이타타입": "TIMESTAMPTZ", "PK여부": ""},
        ]
    },
    "influencer": {
        "테이블명": "인플루언서 데이터",
        "columns": [
            {"컬럼": "id", "컬럼명": "고유 ID", "데이타타입": "INTEGER", "PK여부": "Y"},
            {"컬럼": "influencer_id", "컬럼명": "인플루언서 ID", "데이타타입": "VARCHAR", "PK여부": ""},
            {"컬럼": "influencer_name", "컬럼명": "인플루언서 이름", "데이타타입": "VARCHAR", "PK여부": ""},
            {"컬럼": "platform", "컬럼명": "플랫폼", "데이타타입": "VARCHAR", "PK여부": ""},
            {"컬럼": "content", "컬럼명": "콘텐츠", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "p_coin_name", "컬럼명": "코인 이름", "데이타타입": "VARCHAR", "PK여부": ""},
            {"컬럼": "p_sentiment_score", "컬럼명": "감정 점수", "데이타타입": "DOUBLE PRECISION", "PK여부": ""},
            {"컬럼": "retweet_count", "컬럼명": "리트윗 수", "데이타타입": "INTEGER", "PK여부": ""},
            {"컬럼": "engagement", "컬럼명": "참여도", "데이타타입": "INTEGER", "PK여부": ""},
            {"컬럼": "p_reply_count", "컬럼명": "답글 수", "데이타타입": "INTEGER", "PK여부": ""},
            {"컬럼": "p_like_count", "컬럼명": "좋아요 수", "데이타타입": "INTEGER", "PK여부": ""},
            {"컬럼": "p_quote_count", "컬럼명": "인용 수", "데이타타입": "INTEGER", "PK여부": ""},
            {"컬럼": "post_date", "컬럼명": "게시일", "데이타타입": "TIMESTAMP", "PK여부": ""},
        ]
    },
    "internal_transactions": {
        "테이블명": "내부 거래",
        "columns": [
            {"컬럼": "id", "컬럼명": "고유 ID", "데이타타입": "BIGINT", "PK여부": "Y"},
            {"컬럼": "tx_hash", "컬럼명": "거래 해시", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "trace_id", "컬럼명": "트레이스 ID", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "block_number", "컬럼명": "블록 번호", "데이타타입": "BIGINT", "PK여부": ""},
            {"컬럼": "block_timestamp", "컬럼명": "블록 타임스탬프", "데이타타입": "TIMESTAMPTZ", "PK여부": ""},
            {"컬럼": "from_address", "컬럼명": "보낸 주소", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "to_address", "컬럼명": "받는 주소", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "contract_address", "컬럼명": "컨트랙트 주소", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "value_eth", "컬럼명": "ETH 가치", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "value_usd", "컬럼명": "USD 가치", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "transaction_type", "컬럼명": "거래 유형", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "is_error", "컬럼명": "오류 여부", "데이타타입": "BOOLEAN", "PK여부": ""},
            {"컬럼": "input_data", "컬럼명": "입력 데이터", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "gas", "컬럼명": "가스", "데이타타입": "BIGINT", "PK여부": ""},
            {"컬럼": "gas_used", "컬럼명": "사용된 가스", "데이타타입": "BIGINT", "PK여부": ""},
            {"컬럼": "created_at", "컬럼명": "생성일시", "데이타타입": "TIMESTAMPTZ", "PK여부": ""},
            {"컬럼": "updated_at", "컬럼명": "수정일시", "데이타타입": "TIMESTAMPTZ", "PK여부": ""},
            {"컬럼": "chain", "컬럼명": "체인", "데이타타입": "VARCHAR(50)", "PK여부": ""},
            {"컬럼": "method_id", "컬럼명": "메서드 ID", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "function_name", "컬럼명": "함수 이름", "데이타타입": "VARCHAR", "PK여부": ""},
        ]
    },
    "market_cap_data": {
        "테이블명": "시가총액 데이터",
        "columns": [
            {"컬럼": "id", "컬럼명": "고유 ID", "데이타타입": "UUID", "PK여부": "Y"},
            {"컬럼": "crypto_id", "컬럼명": "암호화폐 ID", "데이타타입": "UUID", "PK여부": ""},
            {"컬럼": "timestamp", "컬럼명": "타임스탬프", "데이타타입": "TIMESTAMPTZ", "PK여부": ""},
            {"컬럼": "market_cap", "컬럼명": "시가총액", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "market_cap_rank", "컬럼명": "시가총액 순위", "데이타타입": "INTEGER", "PK여부": ""},
            {"컬럼": "fully_diluted_market_cap", "컬럼명": "완전 희석 시가총액", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "circulating_supply", "컬럼명": "유통 공급량", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "total_supply", "컬럼명": "총 공급량", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "max_supply", "컬럼명": "최대 공급량", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "market_cap_dominance", "컬럼명": "시가총액 점유율", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "ath_price", "컬럼명": "역대 최고가", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "ath_date", "컬럼명": "역대 최고가 날짜", "데이타타입": "TIMESTAMPTZ", "PK여부": ""},
            {"컬럼": "ath_change_percentage", "컬럼명": "역대 최고가 변동률", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "atl_price", "컬럼명": "역대 최저가", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "atl_date", "컬럼명": "역대 최저가 날짜", "데이타타입": "TIMESTAMPTZ", "PK여부": ""},
            {"컬럼": "atl_change_percentage", "컬럼명": "역대 최저가 변동률", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "data_source", "컬럼명": "데이터 출처", "데이타타입": "VARCHAR", "PK여부": ""},
            {"컬럼": "raw_data", "컬럼명": "원시 데이터", "데이타타입": "JSONB", "PK여부": ""},
            {"컬럼": "created_at", "컬럼명": "생성일시", "데이타타입": "TIMESTAMPTZ", "PK여부": ""},
        ]
    },
    "market_data_daily": {
        "테이블명": "일일 시장 데이터",
        "columns": [
            {"컬럼": "id", "컬럼명": "고유 ID", "데이타타입": "UUID", "PK여부": "Y"},
            {"컬럼": "crypto_id", "컬럼명": "암호화폐 ID", "데이타타입": "UUID", "PK여부": ""},
            {"컬럼": "date", "컬럼명": "날짜", "데이타타입": "DATE", "PK여부": ""},
            {"컬럼": "open_price", "컬럼명": "시작가", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "high_price", "컬럼명": "고가", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "low_price", "컬럼명": "저가", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "close_price", "컬럼명": "종가", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "volume", "컬럼명": "거래량", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "quote_volume", "컬럼명": "기준 통화 거래량", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "price_change_24h", "컬럼명": "24시간 가격 변동", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "price_change_percent_24h", "컬럼명": "24시간 가격 변동률", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "weighted_avg_price", "컬럼명": "가중 평균 가격", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "prev_close_price", "컬럼명": "이전 종가", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "last_price", "컬럼명": "최종 가격", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "bid_price", "컬럼명": "매수 호가", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "ask_price", "컬럼명": "매도 호가", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "trade_count", "컬럼명": "거래 수", "데이타타입": "INTEGER", "PK여부": ""},
            {"컬럼": "first_trade_id", "컬럼명": "첫 거래 ID", "데이타타입": "BIGINT", "PK여부": ""},
            {"컬럼": "last_trade_id", "컬럼명": "마지막 거래 ID", "데이타타입": "BIGINT", "PK여부": ""},
            {"컬럼": "open_time", "컬럼명": "시작 시간", "데이타타입": "TIMESTAMPTZ", "PK여부": ""},
            {"컬럼": "close_time", "컬럼명": "종료 시간", "데이타타입": "TIMESTAMPTZ", "PK여부": ""},
            {"컬럼": "data_source", "컬럼명": "데이터 출처", "데이타타입": "VARCHAR", "PK여부": ""},
            {"컬럼": "raw_data", "컬럼명": "원시 데이터", "데이타타입": "JSONB", "PK여부": ""},
            {"컬럼": "created_at", "컬럼명": "생성일시", "데이타타입": "TIMESTAMPTZ", "PK여부": ""},
        ]
    },
    "news_sentiment": {
        "테이블명": "뉴스 감정 분석",
        "columns": [
            {"컬럼": "id", "컬럼명": "고유 ID", "데이타타입": "UUID", "PK여부": "Y"},
            {"컬럼": "crypto_id", "컬럼명": "암호화폐 ID", "데이타타입": "UUID", "PK여부": ""},
            {"컬럼": "timestamp", "컬럼명": "타임스탬프", "데이타타입": "TIMESTAMPTZ", "PK여부": ""},
            {"컬럼": "news_count", "컬럼명": "뉴스 수", "데이타타입": "INTEGER", "PK여부": ""},
            {"컬럼": "news_sources", "컬럼명": "뉴스 출처", "데이타타입": "ARRAY", "PK여부": ""},
            {"컬럼": "sentiment_score", "컬럼명": "감정 점수", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "sentiment_positive", "컬럼명": "긍정 감정 수", "데이타타입": "INTEGER", "PK여부": ""},
            {"컬럼": "sentiment_neutral", "컬럼명": "중립 감정 수", "데이타타입": "INTEGER", "PK여부": ""},
            {"컬럼": "sentiment_negative", "컬럼명": "부정 감정 수", "데이타타입": "INTEGER", "PK여부": ""},
            {"컬럼": "trending_score", "컬럼명": "트렌딩 점수", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "data_source", "컬럼명": "데이터 출처", "데이타타입": "VARCHAR", "PK여부": ""},
            {"컬럼": "raw_data", "컬럼명": "원시 데이터", "데이타타입": "JSONB", "PK여부": ""},
            {"컬럼": "created_at", "컬럼명": "생성일시", "데이타타입": "TIMESTAMPTZ", "PK여부": ""},
        ]
    },
    "prediction_accuracy": {
        "테이블명": "예측 정확도",
        "columns": [
            {"컬럼": "id", "컬럼명": "고유 ID", "데이타타입": "UUID", "PK여부": "Y"},
            {"컬럼": "analyst_id", "컬럼명": "분석가 ID", "데이타타입": "UUID", "PK여부": ""},
            {"컬럼": "target_id", "컬럼명": "대상 ID", "데이타타입": "UUID", "PK여부": ""},
            {"컬럼": "symbol", "컬럼명": "심볼", "데이타타입": "VARCHAR", "PK여부": ""},
            {"컬럼": "predicted_price", "컬럼명": "예측 가격", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "actual_price", "컬럼명": "실제 가격", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "price_accuracy", "컬럼명": "가격 정확도", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "direction_accuracy", "컬럼명": "방향 정확도", "데이타타입": "BOOLEAN", "PK여부": ""},
            {"컬럼": "timeframe", "컬럼명": "시간 프레임", "데이타타입": "VARCHAR", "PK여부": ""},
            {"컬럼": "prediction_date", "컬럼명": "예측 날짜", "데이타타입": "TIMESTAMPTZ", "PK여부": ""},
            {"컬럼": "evaluation_date", "컬럼명": "평가 날짜", "데이타타입": "TIMESTAMPTZ", "PK여부": ""},
            {"컬럼": "is_evaluated", "컬럼명": "평가 완료 여부", "데이타타입": "BOOLEAN", "PK여부": ""},
            {"컬럼": "created_at", "컬럼명": "생성일시", "데이타타입": "TIMESTAMPTZ", "PK여부": ""},
        ]
    },
    "price_history": {
        "테이블명": "가격 이력",
        "columns": [
            {"컬럼": "id", "컬럼명": "고유 ID", "데이타타입": "UUID", "PK여부": "Y"},
            {"컬럼": "crypto_id", "컬럼명": "암호화폐 ID", "데이타타입": "UUID", "PK여부": ""},
            {"컬럼": "timestamp", "컬럼명": "타임스탬프", "데이타타입": "TIMESTAMPTZ", "PK여부": ""},
            {"컬럼": "open_price", "컬럼명": "시작가", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "high_price", "컬럼명": "고가", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "low_price", "컬럼명": "저가", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "close_price", "컬럼명": "종가", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "volume", "컬럼명": "거래량", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "quote_volume", "컬럼명": "기준 통화 거래량", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "trade_count", "컬럼명": "거래 수", "데이타타입": "INTEGER", "PK여부": ""},
            {"컬럼": "taker_buy_volume", "컬럼명": "매수 거래량", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "taker_buy_quote_volume", "컬럼명": "매수 기준 통화 거래량", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "data_source", "컬럼명": "데이터 출처", "데이타타입": "VARCHAR", "PK여부": ""},
            {"컬럼": "raw_data", "컬럼명": "원시 데이터", "데이타타입": "JSONB", "PK여부": ""},
            {"컬럼": "created_at", "컬럼명": "생성일시", "데이타타입": "TIMESTAMPTZ", "PK여부": ""},
        ]
    },
    "reddit_sentiment": {
        "테이블명": "레딧 감정 분석",
        "columns": [
            {"컬럼": "id", "컬럼명": "고유 ID", "데이타타입": "UUID", "PK여부": "Y"},
            {"컬럼": "crypto_id", "컬럼명": "암호화폐 ID", "데이타타입": "UUID", "PK여부": ""},
            {"컬럼": "timestamp", "컬럼명": "타임스탬프", "데이타타입": "TIMESTAMPTZ", "PK여부": ""},
            {"컬럼": "total_mentions", "컬럼명": "총 언급 수", "데이타타입": "INTEGER", "PK여부": ""},
            {"컬럼": "positive_mentions", "컬럼명": "긍정 언급 수", "데이타타입": "INTEGER", "PK여부": ""},
            {"컬럼": "negative_mentions", "컬럼명": "부정 언급 수", "데이타타입": "INTEGER", "PK여부": ""},
            {"컬럼": "neutral_mentions", "컬럼명": "중립 언급 수", "데이타타입": "INTEGER", "PK여부": ""},
            {"컬럼": "subreddit_breakdown", "컬럼명": "서브레딧 분석", "데이타타입": "JSONB", "PK여부": ""},
            {"컬럼": "sentiment_score", "컬럼명": "감정 점수", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "community_interest", "컬럼명": "커뮤니티 관심도", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "data_source", "컬럼명": "데이터 출처", "데이타타입": "VARCHAR", "PK여부": ""},
            {"컬럼": "raw_data", "컬럼명": "원시 데이터", "데이타타입": "JSONB", "PK여부": ""},
            {"컬럼": "created_at", "컬럼명": "생성일시", "데이타타입": "TIMESTAMPTZ", "PK여부": ""},
        ]
    },
    "social_data": {
        "테이블명": "소셜 데이터",
        "columns": [
            {"컬럼": "id", "컬럼명": "고유 ID", "데이타타입": "UUID", "PK여부": "Y"},
            {"컬럼": "crypto_id", "컬럼명": "암호화폐 ID", "데이타타입": "UUID", "PK여부": ""},
            {"컬럼": "timestamp", "컬럼명": "타임스탬프", "데이타타입": "TIMESTAMPTZ", "PK여부": ""},
            {"컬럼": "twitter_followers", "컬럼명": "트위터 팔로워 수", "데이타타입": "INTEGER", "PK여부": ""},
            {"컬럼": "twitter_following", "컬럼명": "트위터 팔로잉 수", "데이타타입": "INTEGER", "PK여부": ""},
            {"컬럼": "twitter_lists", "컬럼명": "트위터 리스트 수", "데이타타입": "INTEGER", "PK여부": ""},
            {"컬럼": "twitter_favourites", "컬럼명": "트위터 좋아요 수", "데이타타입": "INTEGER", "PK여부": ""},
            {"컬럼": "twitter_statuses", "컬럼명": "트위터 상태 수", "데이타타입": "INTEGER", "PK여부": ""},
            {"컬럼": "reddit_subscribers", "컬럼명": "레딧 구독자 수", "데이타타입": "INTEGER", "PK여부": ""},
            {"컬럼": "reddit_active_users", "컬럼명": "레딧 활성 사용자 수", "데이타타입": "INTEGER", "PK여부": ""},
            {"컬럼": "reddit_posts_per_hour", "컬럼명": "레딧 시간당 게시물 수", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "reddit_comments_per_hour", "컬럼명": "레딧 시간당 댓글 수", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "community_score", "컬럼명": "커뮤니티 점수", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "public_interest_score", "컬럼명": "공개 관심도 점수", "데이타타입": "NUMERIC", "PK여부": ""},
            {"컬럼": "data_source", "컬럼명": "데이터 출처", "데이타타입": "VARCHAR", "PK여부": ""},
            {"컬럼": "raw_data", "컬럼명": "원시 데이터", "데이타타입": "JSONB", "PK여부": ""},
            {"컬럼": "created_at", "컬럼명": "생성일시", "데이타타입": "TIMESTAMPTZ", "PK여부": ""},
        ]
    },
    "whale_address": {
        "테이블명": "고래 지갑 주소",
        "columns": [
            {"컬럼": "id", "컬럼명": "고유 ID", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "chain_type", "컬럼명": "체인 유형", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "address", "컬럼명": "주소", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "name_tag", "컬럼명": "이름 태그", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "balance", "컬럼명": "잔액", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "percentage", "컬럼명": "비율", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "txn_count", "컬럼명": "거래 수", "데이타타입": "TEXT", "PK여부": ""},
        ]
    },
    "whale_transactions": {
        "테이블명": "고래 거래",
        "columns": [
            {"컬럼": "id", "컬럼명": "고유 ID", "데이타타입": "TEXT", "PK여부": "Y"},
            {"컬럼": "tx_hash", "컬럼명": "거래 해시", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "block_number", "컬럼명": "블록 번호", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "block_timestamp", "컬럼명": "블록 타임스탬프", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "from_address", "컬럼명": "보낸 주소", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "to_address", "컬럼명": "받는 주소", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "from_label", "컬럼명": "보낸 주소 라벨", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "to_label", "컬럼명": "받는 주소 라벨", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "coin_symbol", "컬럼명": "코인 심볼", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "amount", "컬럼명": "금액", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "amount_usd", "컬럼명": "USD 금액", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "transaction_type", "컬럼명": "거래 유형", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "is_whale", "컬럼명": "고래 거래 여부", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "whale_category", "컬럼명": "고래 카테고리", "데이타타입": "VARCHAR", "PK여부": ""},
            {"컬럼": "gas_fee", "컬럼명": "가스 비용", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "market_impact_score", "컬럼명": "시장 영향 점수", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "alert_sent", "컬럼명": "알림 전송 여부", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "created_at", "컬럼명": "생성일시", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "gas_used", "컬럼명": "사용된 가스", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "gas_price", "컬럼명": "가스 가격", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "gas_fee_eth", "컬럼명": "ETH 가스 비용", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "gas_fee_usd", "컬럼명": "USD 가스 비용", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "transaction_status", "컬럼명": "거래 상태", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "contract_address", "컬럼명": "컨트랙트 주소", "데이타타입": "VARCHAR", "PK여부": ""},
            {"컬럼": "input_data", "컬럼명": "입력 데이터", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "is_contract_to_contract", "컬럼명": "컨트랙트 간 거래 여부", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "has_method_id", "컬럼명": "메서드 ID 존재 여부", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "updated_at", "컬럼명": "수정일시", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "chain", "컬럼명": "체인", "데이타타입": "VARCHAR(50)", "PK여부": ""},
            {"컬럼": "method_id", "컬럼명": "메서드 ID", "데이타타입": "TEXT", "PK여부": ""},
            {"컬럼": "function_name", "컬럼명": "함수 이름", "데이타타입": "VARCHAR", "PK여부": ""},
        ]
    },
}

def create_excel_with_formatting():
    """Excel 파일 생성 및 포맷팅"""
    output_file = "Supabase_테이블_스키마.xlsx"
    
    # ExcelWriter로 초기 파일 생성
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for table_name, schema_info in table_schemas.items():
            data = []
            for i, col_info in enumerate(schema_info["columns"]):
                data.append({
                    "순번": i + 1,
                    "영역구분": "Supabase",
                    "시스템": "WhaleTracking",
                    "OWNER": "public",
                    "테이블": table_name,
                    "테이블명": schema_info["테이블명"],
                    "순서": i,
                    "컬럼": col_info["컬럼"],
                    "컬럼명": col_info["컬럼명"],
                    "데이타타입": col_info["데이타타입"],
                    "PK여부": col_info["PK여부"],
                })
            
            df = pd.DataFrame(data)
            df.to_excel(writer, sheet_name=table_name, index=False)
    
    # 포맷팅 적용
    wb = load_workbook(output_file)
    
    # 헤더 스타일 정의
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # 연한 파란색
    header_font = Font(bold=True, color="000000")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 헤더 행 스타일 적용
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # 데이터 행에 테두리 적용
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
        
        # 컬럼 너비 자동 조정
        column_widths = {
            'A': 8,   # 순번
            'B': 12,  # 영역구분
            'C': 15,  # 시스템
            'D': 10,  # OWNER
            'E': 25,  # 테이블
            'F': 25,  # 테이블명
            'G': 8,   # 순서
            'H': 25,  # 컬럼
            'I': 30,  # 컬럼명
            'J': 25,  # 데이타타입
            'K': 10,  # PK여부
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 행 높이 조정
        ws.row_dimensions[1].height = 25  # 헤더 행 높이
    
    wb.save(output_file)
    print(f"✅ Excel 파일 생성 완료: {output_file}")
    print(f"   총 {len(table_schemas)}개의 테이블이 {len(table_schemas)}개의 탭으로 구성되었습니다.")

if __name__ == '__main__':
    create_excel_with_formatting()



